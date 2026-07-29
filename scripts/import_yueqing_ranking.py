#!/usr/bin/env python3
"""Import a Yueqing cumulative export ranking workbook into stable static JSON."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


RANGE_PATTERN = re.compile(r"^(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)万美元$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--period-id", default="2026-ytd-06")
    parser.add_argument("--period-label", default="2026年1—6月")
    parser.add_argument("--as-of", default="2026-06-30")
    return parser.parse_args()


def load_rows(path: Path) -> list[dict]:
    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = []
    for excel_row, (raw,) in enumerate(
        sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True),
        start=2,
    ):
        if raw in (None, ""):
            continue
        fields = [value.strip() for value in next(csv.reader([str(raw)]))]
        if len(fields) != 3:
            raise ValueError(f"Excel row {excel_row} does not contain three CSV fields")
        rank_text, company_name, band_label = fields
        rows.append(
            {
                "source_row": excel_row,
                "rank": int(rank_text),
                "company_name": company_name,
                "band_label": band_label,
            }
        )
    return rows


def build_dataset(rows: list[dict], args: argparse.Namespace) -> tuple[list[dict], dict]:
    ranks = [row["rank"] for row in rows]
    if ranks != list(range(1, len(rows) + 1)):
        raise ValueError("Ranks must be continuous and sorted from 1")

    name_counts = Counter(row["company_name"] for row in rows)
    company_ids: dict[str, str] = {}
    companies = []
    for row in rows:
        name = row["company_name"]
        if name in company_ids:
            continue
        company_id = f"YQ{len(company_ids) + 1:06d}"
        company_ids[name] = company_id
        companies.append(
            {
                "company_id": company_id,
                "company_name": name,
                "aliases": [],
                "needs_review": name_counts[name] > 1,
            }
        )

    band_labels = []
    for row in rows:
        if row["band_label"] not in band_labels:
            band_labels.append(row["band_label"])

    bands = []
    for index, label in enumerate(band_labels, start=1):
        match = RANGE_PATTERN.fullmatch(label)
        bands.append(
            {
                "id": f"b{index:02d}",
                "label": label,
                "min_usd_10k": float(match.group(1)) if match else None,
                "max_usd_10k": float(match.group(2)) if match else None,
                "count": sum(row["band_label"] == label for row in rows),
            }
        )
    band_ids = {band["label"]: band["id"] for band in bands}

    records = [
        {
            "rank": row["rank"],
            "company_id": company_ids[row["company_name"]],
            "company_name": row["company_name"],
            "band_id": band_ids[row["band_label"]],
            "band_label": row["band_label"],
            "source_row": row["source_row"],
            "source_duplicate": name_counts[row["company_name"]] > 1,
        }
        for row in rows
    ]
    data = {
        "period": {
            "id": args.period_id,
            "label": args.period_label,
            "type": "year_to_date",
            "as_of": args.as_of,
        },
        "source_note": "排名来自用户提供的区间数据；仅展示累计排名与出口额区间，不推算精确出口额。",
        "bands": bands,
        "records": records,
    }
    return companies, data


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )


def main() -> None:
    args = parse_args()
    rows = load_rows(args.workbook)
    companies, data = build_dataset(rows, args)
    args.output.mkdir(parents=True, exist_ok=True)
    write_json(args.output / "companies.json", companies)
    write_json(args.output / f"{args.period_id}.json", data)
    write_json(
        args.output / "periods.json",
        {
            "latest": args.period_id,
            "periods": [
                {
                    **data["period"],
                    "file": f"{args.period_id}.json",
                    "record_count": len(data["records"]),
                }
            ],
        },
    )
    print(f"Imported {len(data['records'])} rows and {len(companies)} companies")


if __name__ == "__main__":
    main()
